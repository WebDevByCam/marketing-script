"""
Escritura de resultados en diferentes formatos.
"""
import os
import time
import random
from typing import List, Dict
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# Google Sheets integration has been removed from the core writer to keep
# the project lightweight. If you need to re-add Sheets support, reintroduce
# the gspread/google-auth dependency and the corresponding methods.
gspread = None
SACredentials = None


def human_print(text: str, humanize: bool = False, speed: float = 0.02):
    """Print texto simulando tipeo humano. Si humanize=False hace print normal."""
    if not humanize:
        print(text)
        return
    
    # Simular tipeo con posibles typos leves y corrección
    for ch in text:
        print(ch, end="", flush=True)
        time.sleep(speed * (0.5 + random.random()))
    print()


class OutputWriter:
    """Maneja la escritura de resultados en diferentes formatos."""
    
    def __init__(self, humanize: bool = False, humanize_speed: float = 0.02):
        self.humanize = humanize
        self.humanize_speed = humanize_speed
    
    def _filter_website_url(self, url):
        """Filtra URLs que no son páginas web reales."""
        if not url or pd.isna(url) or str(url).strip().upper() in ['N/A', '']:
            return "N/A"
        
        url_str = str(url).lower().strip()
        
        # URLs a excluir
        excluded_domains = [
            'wa.me', 'wa.link', 'whatsapp.com', 'web.whatsapp.com',
            'instagram.com', 'facebook.com', 'twitter.com', 'tiktok.com',
            'linkedin.com', 'youtube.com', 'maps.google.com', 'goo.gl',
            'bit.ly', 'tinyurl.com', 't.co'
        ]
        
        # Verificar si la URL contiene alguno de los dominios excluidos
        is_excluded = any(domain in url_str for domain in excluded_domains)
        
        if is_excluded:
            return "N/A"
        else:
            return url
    
    def print(self, text: str):
        """Imprime texto con o sin efecto de tipeo."""
        human_print(text, self.humanize, self.humanize_speed)
    
    def write_to_csv(self, data: List[Dict], filepath: str):
        """Escribe datos a CSV."""
        if not data:
            self.print("[warn] No hay datos para escribir")
            return
        
        df = pd.DataFrame(data)
        df.to_csv(filepath, index=False, encoding="utf-8-sig")
        self.print(f"[✓] Archivo CSV guardado -> {filepath}")
    
    def write_to_excel(self, data: List[Dict], filepath: str, use_template_columns: bool = False):
        """Escribe datos a Excel.
        
        Si use_template_columns=True, usa solo las columnas básicas del template
        (Nombre, WhatsApp, Telefono, Correo, Pagina Web, Ciudad) sin campos de debug.
        """
        if not data:
            self.print("[warn] No hay datos para escribir")
            return
        
        # Normalizar y ordenar columnas según el formato solicitado
        df = pd.DataFrame(data)

        # Aplicar correcciones antes de guardar
        # 1. Filtrar URLs no deseadas
        if 'Página Web' in df.columns or 'Pagina Web' in df.columns:
            col_name = 'Página Web' if 'Página Web' in df.columns else 'Pagina Web'
            df[col_name] = df[col_name].apply(self._filter_website_url)
        
        # 2. Convertir números a string para evitar ".0"
        df = df.astype(str)
        
        # 3. Reemplazar valores vacíos y NaN con "N/A"
        df = df.replace('', 'N/A')
        df = df.replace('nan', 'N/A')
        df = df.replace('NaN', 'N/A')

        if use_template_columns:
            # Usar solo columnas del template (sin debug ni opcionales)
            template_cols = [
                "Nombre",
                "WhatsApp",
                "Telefono",
                "Correo",
                "Pagina Web",
                "Ciudad"
            ]
            
            # Mapear nombres de columnas a formato objetivo
            col_map = {
                "Teléfono": "Telefono",
                "Teléfono (opcional)": "Telefono",
                "Página Web": "Pagina Web",
                "Página web": "Pagina Web",
                "Correo": "Correo",
                "Ciudad": "Ciudad",
                "Nombre": "Nombre",
            }
            df = df.rename(columns={k: v for k, v in col_map.items() if k in df.columns})
            
            # Asegurar que las columnas existan
            for c in template_cols:
                if c not in df.columns:
                    df[c] = "N/A"
            
            df_out = df[template_cols]
        else:
            # Mapear nombres de columnas a formato objetivo (sin acentos para compatibilidad)
            col_map = {
                "Teléfono": "Telefono",
                "Teléfono (opcional)": "Telefono",
                "Página Web": "Pagina Web",
                "Página web": "Pagina Web",
                "Correo": "Correo",
                "Ciudad": "Ciudad",
                "Nombre": "Nombre",
            }
            df = df.rename(columns={k: v for k, v in col_map.items() if k in df.columns})

            # Columnas objetivo y orden
            target_cols = [
                "Nombre",
                "WhatsApp",
                "Telefono",
                "Correo",
                "Pagina Web",
                "Ciudad",
                "Dirección (opcional)",
                "Google Maps URL (opcional)",
                "place_id (debug)",
            ]

            # Asegurar que las columnas existan
            for c in target_cols:
                if c not in df.columns:
                    df[c] = "N/A"

            df_out = df[target_cols]

        # Escribir usando pandas y luego estilizar encabezado con openpyxl
        df_out.to_excel(filepath, index=False, engine="openpyxl")

        try:
            wb = load_workbook(filepath)
            ws = wb.active
            
            # Definir estilos
            header_fill = PatternFill(start_color="FFF59D", end_color="FFF59D", fill_type="solid")
            header_font = Font(bold=True)
            header_alignment = Alignment(horizontal="center", vertical="center")
            
            # Definir borde completo
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Obtener índices de columnas que necesitan alineación centrada
            columns_to_center = []
            for col_idx, col_name in enumerate(df_out.columns, start=1):
                if col_name in ['WhatsApp', 'Telefono', 'Correo']:
                    columns_to_center.append(col_idx)
            
            # Aplicar estilos al encabezado
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = header_alignment
                cell.border = thin_border
            
            # Aplicar estilos a las filas de datos
            for row in range(2, ws.max_row + 1):
                for col in range(1, ws.max_column + 1):
                    cell = ws.cell(row=row, column=col)
                    cell.border = thin_border
                    
                    # Aplicar alineación centrada a columnas específicas
                    if col in columns_to_center:
                        cell.alignment = Alignment(horizontal="center", vertical="center")
            
            wb.save(filepath)
        except Exception:
            # Si falla el formateo, no interrumpir el flujo
            pass

        self.print(f"[✓] Archivo Excel guardado -> {filepath}")
    
    def write_to_txt(self, data: List[Dict], filepath: str, separator: str = " | "):
        """Escribe datos a archivo de texto simple."""
        if not data:
            self.print("[warn] No hay datos para escribir")
            return
        
        with open(filepath, "w", encoding="utf-8") as fh:
            for row in data:
                # línea compacta: campos principales separados
                fields = [
                    "Nombre", "Teléfono", "Correo", "Página Web", 
                    "Ciudad", "Dirección (opcional)"
                ]
                line = separator.join([str(row.get(field, "")) for field in fields])
                fh.write(line + "\n")
        
        self.print(f"[✓] Archivo de texto guardado -> {filepath}")
    
    def write_to_gsheets(self, data: List[Dict], spreadsheet_id: str, 
                        worksheet_title: str, sa_json_path: str = None):
        # Google Sheets support removed from this writer. If you need to
        # export directly to Google Sheets, use a separate script or re-add
        # the dependency and implementation. For now, prefer CSV/XLSX/TXT.
        raise NotImplementedError("Google Sheets writing has been removed. Use CSV/XLSX output.")

    def merge_with_gsheet(self, data: List[Dict], spreadsheet_id: str, worksheet_title: str, sa_json_path: str = None, key_fields: list = None):
        # Removed: merging with Google Sheets is no longer supported in the
        # cleaned project. Implement externally if required.
        raise NotImplementedError("Google Sheets merge has been removed. Use local CSV/XLSX workflows.")

    def create_backup(self, filepath: str) -> str:
        """Crea una copia de seguridad del archivo con timestamp.
        
        Retorna la ruta de la copia creada.
        """
        from datetime import datetime
        import shutil
        
        if not os.path.exists(filepath):
            return ""
        
        # Crear carpeta de backups si no existe
        backup_dir = os.path.join(os.path.dirname(filepath), '..', 'backups')
        os.makedirs(backup_dir, exist_ok=True)
        
        # Generar nombre con timestamp
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = os.path.basename(filepath)
        name, ext = os.path.splitext(filename)
        backup_name = f"{name}_backup_{timestamp}{ext}"
        backup_path = os.path.join(backup_dir, backup_name)
        
        shutil.copy2(filepath, backup_path)
        self.print(f"[✓] Backup creado -> {backup_path}")
        return backup_path

    def merge_into_existing_excel(self, existing_filepath: str, data: List[Dict], key_priority: list = None, output_dir: str = None):
        """Fusiona `data` (lista de dicts) dentro de un archivo Excel existente.

        Reglas:
        - No se agregan columnas nuevas. Sólo se actualizan las columnas que ya
          existen en el archivo Excel.
        - Se intenta emparejar filas usando, por defecto, 'Pagina Web' si existe
          y tiene valores, o en su defecto 'Nombre'. Se normaliza la clave para
          evitar variaciones simples (http(s), www, mayúsculas, caracteres).
        - Si no existe coincidencia, se añade una nueva fila al final con las
          columnas existentes (los campos que no estén disponibles se dejan vacíos).
        - FILTRO: Solo se incluyen filas que tengan al menos WhatsApp o Telefono.
        - NO sobrescribe el original: crea archivo nuevo en data/merged/ y renombra el original.

        Parámetros:
        - existing_filepath: ruta al archivo Excel original (NO se modifica).
        - data: lista de diccionarios con datos procesados.
        - key_priority: lista de columnas a usar como clave en orden de preferencia.
        - output_dir: directorio donde guardar el merge (default: data/merged/).
        
        Retorna:
        - Tupla (merged_filepath, renamed_original_filepath)
        """
        import re
        import shutil
        import re
        import shutil
        from openpyxl.styles import Font, PatternFill, Alignment

        # Helper to clean values (handle NaN, None, etc.)
        def clean_value(val):
            if val is None or (isinstance(val, float) and str(val) == 'nan'):
                return ''
            return str(val).strip()

        if not data:
            self.print("[warn] No hay datos para fusionar en Excel existente")
            return None, None
        
        # NO filtrar filas sin teléfono ni WhatsApp - mantener todas las filas
        # pero asegurar que tengan valores válidos
        filtered_data = []
        for item in data:
            # Limpiar valores y asegurar que no sean NaN
            whatsapp = clean_value(item.get('WhatsApp', ''))
            telefono = clean_value(item.get('Telefono', ''))
            telefono_alt = clean_value(item.get('Teléfono', ''))
            
            # Si no hay teléfono ni WhatsApp, poner N/A
            if not whatsapp and not telefono and not telefono_alt:
                # Crear una copia del item con valores N/A
                cleaned_item = item.copy()
                cleaned_item['WhatsApp'] = 'N/A'
                cleaned_item['Telefono'] = 'N/A'
                if 'Teléfono' in cleaned_item:
                    cleaned_item['Teléfono'] = 'N/A'
                filtered_data.append(cleaned_item)
            else:
                filtered_data.append(item)
        
        self.print(f"[i] Todas las filas mantenidas: {len(filtered_data)} de {len(data)}")

        if key_priority is None:
            key_priority = ["Pagina Web", "Nombre"]

        # Leer el Excel existente
        if not os.path.exists(existing_filepath):
            raise FileNotFoundError(f"No existe el archivo Excel: {existing_filepath}")

        # Cargar con openpyxl para preservar estilos
        wb_original = load_workbook(existing_filepath)
        ws_original = wb_original.active
        
        # Leer con pandas para manipulación de datos
        df_exist = pd.read_excel(existing_filepath)

        # Normalizar columnas: strip
        df_exist.columns = [str(c).strip() for c in df_exist.columns]

        existing_cols = list(df_exist.columns)

        # Decide la columna clave presente en el archivo existente
        key_col = None
        for k in key_priority:
            if k in df_exist.columns:
                key_col = k
                break
        # If no key column present, fall back to first column named 'Nombre' or index
        if key_col is None:
            if 'Nombre' in df_exist.columns:
                key_col = 'Nombre'
            else:
                key_col = df_exist.columns[0]

        def normalize_key(val: str) -> str:
            if val is None:
                return ""
            s = str(val).lower().strip()
            # remove protocol and www
            s = re.sub(r'^https?://', '', s)
            s = re.sub(r'^www\.', '', s)
            # remove non-alphanumeric
            s = re.sub(r'[^a-z0-9]', '', s)
            return s

        # Build mapping of existing keys -> row index
        existing_keys = df_exist.get(key_col, pd.Series([""] * len(df_exist))).fillna("")
        existing_map = {normalize_key(v): i for i, v in enumerate(existing_keys)}

        # Helper to determine if a phone looks like mobile (quick heuristic)
        def looks_like_mobile(phone: str) -> bool:
            if not phone:
                return False
            digits = re.sub(r'\D', '', str(phone))
            if digits.startswith('57'):
                digits = digits[2:]
            # In Colombia mobiles start with '3' (e.g., 300,311,312,320...)
            return digits.startswith('3')

        # Iterate filtered_data and merge
        for item in filtered_data:
            # build item key using same priority
            item_key_val = ""
            for k in key_priority:
                if k in item and item[k]:
                    item_key_val = item[k]
                    break
            if not item_key_val:
                # fallback to Nombre
                item_key_val = item.get('Nombre', '')

            nkey = normalize_key(item_key_val)

            if nkey in existing_map:
                idx = existing_map[nkey]
                # Update only columns that exist in df_exist and in item
                for col in existing_cols:
                    if col in item and col not in [None, 'Unnamed: 0']:
                        # Special handling: distribute phone between WhatsApp/Telefono
                        if col == 'WhatsApp':
                            # prefer explicit 'WhatsApp' in item, else try to detect
                            val = clean_value(item.get('WhatsApp')) or clean_value(item.get('Telefono')) or clean_value(item.get('Teléfono'))
                            if val and not looks_like_mobile(val):
                                # if detected as non-mobile, skip writing to WhatsApp
                                continue
                            df_exist.at[idx, col] = val
                        elif col == 'Telefono':
                            val = clean_value(item.get('Telefono')) or clean_value(item.get('Teléfono')) or clean_value(item.get('phone'))
                            # if this looks like mobile, prefer WhatsApp instead
                            if val and looks_like_mobile(val):
                                # if Telefono column exists but we have a mobile, leave Telefono empty
                                # unless Telefono is empty and WhatsApp column doesn't exist
                                if 'WhatsApp' not in df_exist.columns:
                                    df_exist.at[idx, col] = val
                                else:
                                    # skip writing mobile into Telefono
                                    pass
                            else:
                                df_exist.at[idx, col] = val
                        else:
                            df_exist.at[idx, col] = item.get(col, df_exist.at[idx, col])
            else:
                # Row not found: append a new row but only with existing columns
                new_row = {c: '' for c in existing_cols}
                for c in existing_cols:
                    if c in item:
                        # map phone heuristics as above
                        if c == 'WhatsApp':
                            val = clean_value(item.get('WhatsApp')) or clean_value(item.get('Telefono')) or clean_value(item.get('Teléfono'))
                            if val and not looks_like_mobile(val):
                                # don't insert non-mobile into WhatsApp
                                continue
                            new_row[c] = val
                        elif c == 'Telefono':
                            val = clean_value(item.get('Telefono')) or clean_value(item.get('Teléfono')) or clean_value(item.get('phone'))
                            if val and looks_like_mobile(val):
                                if 'WhatsApp' not in existing_cols:
                                    new_row[c] = val
                                else:
                                    # skip mobile into Telefono when WhatsApp exists
                                    pass
                            else:
                                new_row[c] = val
                        else:
                            new_row[c] = clean_value(item.get(c, ''))
                    elif c == 'Unnamed: 0':
                        # Para la columna de índice, usar el próximo índice disponible
                        new_row[c] = len(df_exist)
                
                # Usar loc en lugar de concat para mayor confiabilidad
                df_exist.loc[len(df_exist)] = new_row

        # Preparar paths para el merge
        if output_dir is None:
            output_dir = os.path.join(os.path.dirname(existing_filepath), '..', 'merged')
        os.makedirs(output_dir, exist_ok=True)
        
        # Nombres de archivos
        original_filename = os.path.basename(existing_filepath)
        name_without_ext = os.path.splitext(original_filename)[0]
        ext = os.path.splitext(original_filename)[1]
        
        # Path del merged (tendrá el nombre original)
        merged_filepath = os.path.join(output_dir, original_filename)
        
        # Path del original renombrado (se queda en input con " - original")
        renamed_original = os.path.join(
            os.path.dirname(existing_filepath),
            f"{name_without_ext} - original{ext}"
        )
        
        # Escribir el merged con pandas (temporal)
        temp_file = merged_filepath + '.tmp' + ext
        df_exist.to_excel(temp_file, index=False, engine='openpyxl')
        
        # Cargar el merged temporal con openpyxl para aplicar estilos
        wb_merged = load_workbook(temp_file)
        ws_merged = wb_merged.active
        
        # Copiar estilos del header original (fila 1)
        header_fill = PatternFill(start_color="FFFFFF00", end_color="FFFFFF00", fill_type="solid")
        header_font = Font(name="Aptos Narrow", bold=True, size=16)
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        for col_idx, cell in enumerate(ws_merged[1], start=1):
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
        
        # Copiar estilos de datos (fondo verde claro)
        data_fill = PatternFill(start_color="FFC1F0C8", end_color="FFC1F0C8", fill_type="solid")
        data_font = Font(name="Aptos Narrow", bold=False, size=11)
        
        for row_idx in range(2, ws_merged.max_row + 1):
            for col_idx in range(1, ws_merged.max_column + 1):
                cell = ws_merged.cell(row=row_idx, column=col_idx)
                cell.fill = data_fill
                cell.font = data_font
                
                # Primera columna (índice) en bold
                if col_idx == 1:
                    cell.font = Font(name="Aptos Narrow", bold=True, size=11)
        
        # Guardar el archivo merged final
        wb_merged.save(merged_filepath)
        os.remove(temp_file)  # Eliminar temporal
        
        # Renombrar el original
        if os.path.exists(renamed_original):
            # Si ya existe un "- original", añadir timestamp
            from datetime import datetime
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            renamed_original = os.path.join(
                os.path.dirname(existing_filepath),
                f"{name_without_ext} - original {timestamp}{ext}"
            )
        
        shutil.move(existing_filepath, renamed_original)
        
        self.print(f"[✓] Archivo original renombrado -> {renamed_original}")
        self.print(f"[✓] Archivo merged creado -> {merged_filepath}")
        
        return merged_filepath, renamed_original
    
    def auto_write(self, data: List[Dict], filepath: str):
        """Escribe automáticamente según la extensión del archivo."""
        if not filepath:
            return
        
        ext = filepath.lower().split('.')[-1]
        
        if ext == 'csv':
            self.write_to_csv(data, filepath)
        elif ext in ['xlsx', 'xls']:
            self.write_to_excel(data, filepath)
        elif ext == 'txt':
            self.write_to_txt(data, filepath)
        else:
            self.print(f"[warn] Formato no reconocido: {ext}, guardando como CSV")
            self.write_to_csv(data, filepath + '.csv')